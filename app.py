import streamlit as st
import pdfplumber
import pandas as pd
import re
from datetime import datetime
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("Erro: A biblioteca 'openpyxl' não está instalada. Certifique-se de incluir 'openpyxl' no seu arquivo requirements.txt.")

# --- 1. CONFIGURAÇÃO E ESTILO ---
st.set_page_config(page_title="Edson Medeiros | Consultoria e Compliance", layout="wide", page_icon="⚖️")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,400;0,500;0,600;0,700;1,400;1,600&family=Inter:wght@300;400;500;600&family=Great+Vibes&display=swap');

/* ═══ VARIÁVEIS ════════════════════════════════════════════════════════════ */
:root {
    --p:     #050810;
    --p2:    #0A0F18;
    --p3:    #111823;
    --p4:    #19222F;
    --g:     #C5A566;
    --g2:    #D4B87A;
    --g3:    #A8883E;
    --c:     #EDE5D4;
    --cm:    rgba(237,229,212,0.55);
    --cl:    rgba(237,229,212,0.20);
    --gl:    rgba(197,165,102,0.15);
    --gm:    rgba(197,165,102,0.08);
    --serif: 'Cormorant Garamond', Georgia, serif;
    --sans:  'Inter', system-ui, sans-serif;
    --r-sm:  10px;
    --r-md:  16px;
    --r-lg:  24px;
    --r-xl:  32px;
}

/* ═══ RESET BASE ══════════════════════════════════════════════════════════ */
html, body, [class*="css"] {
    font-family: var(--sans);
    -webkit-font-smoothing: antialiased;
    text-rendering: optimizeLegibility;
}

/* ═══ FUNDO EM CAMADAS ════════════════════════════════════════════════════ */
.stApp {
    background: var(--p);
    color: var(--c);
    /* Camada 1: grade de pontos */
    background-image:
        radial-gradient(circle, rgba(197,165,102,0.055) 1px, transparent 1px);
    background-size: 28px 28px;
}
/* Camada 2: luz dourada central — injetada via pseudo-elemento no body */
body::before {
    content: '';
    position: fixed;
    top: -30vh; left: 50%;
    transform: translateX(-50%);
    width: 80vw; height: 80vh;
    background: radial-gradient(ellipse,
        rgba(197,165,102,0.055) 0%,
        transparent 65%);
    pointer-events: none;
    z-index: 0;
}

/* ═══ ANIMAÇÕES ═══════════════════════════════════════════════════════════ */
@keyframes fadeUp {
    from { opacity:0; transform:translateY(32px); }
    to   { opacity:1; transform:translateY(0); }
}
@keyframes fadeIn {
    from { opacity:0; }
    to   { opacity:1; }
}
@keyframes shimmerText {
    0%   { background-position: -500px 0; }
    100% { background-position: 500px 0; }
}
@keyframes pulseRing {
    0%,100% { opacity:0.25; transform:scale(1); }
    50%      { opacity:0.7;  transform:scale(1.05); }
}
@keyframes glow {
    0%,100% { opacity:0.3; }
    50%      { opacity:0.85; }
}
@keyframes spin {
    to { transform: rotate(360deg); }
}
@keyframes barFill {
    from { width: 0; }
    to   { width: var(--bar-w, 60%); }
}
@keyframes floatY {
    0%,100% { transform: translateY(0); }
    50%      { transform: translateY(-5px); }
}
@keyframes revealLine {
    from { scaleX: 0; }
    to   { scaleX: 1; }
}

/* ═══ CONTAINER ═══════════════════════════════════════════════════════════ */
.block-container {
    max-width: 1100px !important;
    padding: 0 40px 100px !important;
    position: relative; z-index: 1;
    animation: fadeUp 0.7s cubic-bezier(.22,1,.36,1) both;
}

/* ═══ HEADER ═══════════════════════════════════════════════════════════════ */
.em-header-wrap {
    text-align: center;
    padding: 72px 0 52px;
    position: relative;
    animation: fadeUp 0.8s cubic-bezier(.22,1,.36,1) both;
}
/* Linhas laterais com animação */
.em-header-wrap::before, .em-header-wrap::after {
    content: '';
    position: absolute;
    top: 50%;
    width: calc(50% - 260px);
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(197,165,102,0.4));
    animation: glow 3.5s ease-in-out infinite;
}
.em-header-wrap::before { left: 0; transform: translateY(-50%) scaleX(-1); }
.em-header-wrap::after  { right: 0; transform: translateY(-50%); }

/* Monograma — círculo dourado acima do nome */
.em-monogram {
    display: inline-flex;
    align-items: center; justify-content: center;
    width: 52px; height: 52px;
    border-radius: 50%;
    border: 1px solid rgba(197,165,102,0.35);
    margin-bottom: 16px;
    position: relative;
    animation: floatY 4s ease-in-out infinite;
}
.em-monogram::before {
    content: '';
    position: absolute; inset: -5px;
    border-radius: 50%;
    border: 1px solid rgba(197,165,102,0.12);
    animation: pulseRing 3s ease-in-out infinite;
}
.em-monogram-text {
    font-family: var(--serif);
    font-size: 1.1rem; font-weight: 600;
    color: var(--g); letter-spacing: 3px;
}

.em-eyebrow {
    font-family: var(--sans);
    font-size: 0.68rem; font-weight: 600;
    letter-spacing: 5px; text-transform: uppercase;
    color: rgba(197,165,102,0.42);
    margin-bottom: 12px;
}
.em-name {
    font-family: var(--serif);
    font-size: clamp(3rem, 5.5vw, 4.8rem);
    font-weight: 600; line-height: 1.0;
    letter-spacing: 1px; margin: 0;
    background: linear-gradient(110deg,
        #EDE5D4 15%, #C5A566 40%, #D4B87A 55%, #EDE5D4 75%);
    background-size: 600px 100%;
    background-clip: text; -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    animation: shimmerText 5s linear infinite;
}
.em-subtitle {
    font-family: var(--serif);
    font-size: 0.9rem; font-style: italic;
    color: rgba(197,165,102,0.5);
    letter-spacing: 3.5px; margin-top: 10px;
    text-transform: uppercase;
}
.em-badge-row {
    display: flex; align-items: center; justify-content: center;
    gap: 20px; margin-top: 22px;
}
.em-badge {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 6px 16px;
    border: 1px solid rgba(197,165,102,0.18);
    border-radius: 100px;
    font-size: 0.7rem; font-weight: 600; letter-spacing: 1.5px;
    text-transform: uppercase; color: rgba(197,165,102,0.5);
}
.em-badge-dot {
    width: 5px; height: 5px; border-radius: 50%;
    background: var(--g); opacity: 0.6;
    animation: glow 2s ease-in-out infinite;
}
.em-ornament {
    display: flex; align-items: center; justify-content: center;
    gap: 14px; margin-top: 24px;
}
.em-ornament-line {
    width: 80px; height: 1px;
    background: linear-gradient(90deg, transparent, rgba(197,165,102,0.45));
    animation: glow 3s ease-in-out infinite;
}
.em-ornament-line.rev {
    background: linear-gradient(90deg, rgba(197,165,102,0.45), transparent);
}
.em-ornament-diamond { color: rgba(197,165,102,0.65); font-size: 0.55rem; }

/* ═══ DIVISORES ═══════════════════════════════════════════════════════════ */
.em-divider {
    display: flex; align-items: center; gap: 16px;
    margin: 44px 0 30px;
}
.em-divider-line {
    flex: 1; height: 1px;
    background: linear-gradient(90deg,
        rgba(197,165,102,0.05), rgba(197,165,102,0.2), rgba(197,165,102,0.05));
}
.em-divider-pill {
    display: flex; align-items: center; gap: 8px;
    padding: 5px 16px;
    border: 1px solid rgba(197,165,102,0.18);
    border-radius: 100px;
    background: rgba(197,165,102,0.04);
}
.em-divider-label {
    font-size: 0.68rem; font-weight: 600;
    letter-spacing: 3.5px; text-transform: uppercase;
    color: rgba(197,165,102,0.45); white-space: nowrap;
}
.em-divider-num {
    font-family: var(--serif);
    font-size: 0.75rem; color: rgba(197,165,102,0.35);
}
.em-section-note {
    font-size: 0.88rem; color: rgba(237,229,212,0.3);
    text-align: center; letter-spacing: 0.3px;
    margin: 4px 0 28px; line-height: 1.6;
}

/* ═══ UPLOAD ZONE ══════════════════════════════════════════════════════════ */
.upload-wrap {
    border-radius: var(--r-lg);
    border: 1px solid rgba(197,165,102,0.18);
    background: linear-gradient(145deg,
        rgba(197,165,102,0.04) 0%,
        rgba(10,15,24,0.8) 100%);
    padding: 4px;
    transition: all 0.4s cubic-bezier(.22,1,.36,1);
    animation: fadeUp 0.5s cubic-bezier(.22,1,.36,1) both;
}
.upload-wrap:hover {
    border-color: rgba(197,165,102,0.45);
    box-shadow: 0 0 60px rgba(197,165,102,0.07),
                inset 0 1px 0 rgba(197,165,102,0.08);
}
[data-testid="stFileUploader"] {
    border-radius: var(--r-md) !important;
    border: none !important;
    background: transparent !important;
}
[data-testid="stFileUploadDropzone"] {
    background: transparent !important;
    border: 1px dashed rgba(197,165,102,0.25) !important;
    border-radius: var(--r-md) !important;
    padding: 44px !important;
    transition: all 0.3s ease;
}
[data-testid="stFileUploader"]:hover [data-testid="stFileUploadDropzone"] {
    border-color: rgba(197,165,102,0.5) !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] span,
[data-testid="stFileUploaderDropzoneInstructions"] small {
    color: rgba(197,165,102,0.55) !important;
    font-family: var(--sans) !important;
    font-size: 0.9rem !important; letter-spacing: 0.5px;
}
[data-testid="stFileUploadDropzone"] button {
    background: rgba(197,165,102,0.07) !important;
    border: 1px solid rgba(197,165,102,0.35) !important;
    color: var(--g) !important;
    border-radius: var(--r-sm) !important;
    font-family: var(--sans) !important;
    font-size: 0.66rem !important; font-weight: 600 !important;
    letter-spacing: 2px !important; text-transform: uppercase !important;
    padding: 9px 24px !important;
    transition: all 0.25s ease;
}
[data-testid="stFileUploadDropzone"] button:hover {
    background: rgba(197,165,102,0.14) !important;
    border-color: var(--g) !important;
    transform: translateY(-1px) !important;
}

/* Tutorial de upload */
.upload-tutorial {
    display: flex; gap: 24px; justify-content: center;
    flex-wrap: wrap;
    padding: 20px 0 4px;
}
.upload-step {
    display: flex; align-items: center; gap: 8px;
}
.upload-step-n {
    width: 22px; height: 22px;
    border-radius: 50%; border: 1px solid rgba(197,165,102,0.25);
    display: flex; align-items: center; justify-content: center;
    font-size: 0.6rem; font-weight: 600; color: rgba(197,165,102,0.55);
    flex-shrink: 0;
}
.upload-step-t {
    font-size: 0.82rem; color: rgba(237,229,212,0.35); letter-spacing: 0.3px;
}
.upload-step-arrow {
    font-size: 0.6rem; color: rgba(197,165,102,0.2);
    margin: 0 -12px; align-self: center;
}

/* ═══ METRIC CARDS ═════════════════════════════════════════════════════════ */
.metric-card {
    background: linear-gradient(145deg, var(--p3) 0%, var(--p2) 100%);
    border: 1px solid rgba(197,165,102,0.12);
    border-top: none;
    border-radius: var(--r-md);
    padding: 28px 24px 24px;
    text-align: left;
    position: relative; overflow: hidden;
    transition: all 0.35s cubic-bezier(.22,1,.36,1);
    animation: fadeUp 0.5s cubic-bezier(.22,1,.36,1) both;
}
/* Linha superior colorida */
.metric-card::before {
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, transparent, var(--g), transparent);
    opacity: 0.5;
    transition: opacity 0.3s ease;
}
/* Reflexo de luz interno */
.metric-card::after {
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 50%;
    background: linear-gradient(180deg,
        rgba(197,165,102,0.04) 0%, transparent 100%);
    border-radius: var(--r-md) var(--r-md) 0 0;
    pointer-events: none;
}
.metric-card:hover {
    transform: translateY(-4px) scale(1.01);
    border-color: rgba(197,165,102,0.28);
    box-shadow: 0 20px 60px rgba(0,0,0,0.4),
                0 0 30px rgba(197,165,102,0.07);
}
.metric-card:hover::before { opacity: 1; }

.metric-icon {
    width: 36px; height: 36px;
    border-radius: var(--r-sm);
    border: 1px solid rgba(197,165,102,0.2);
    background: rgba(197,165,102,0.07);
    display: flex; align-items: center; justify-content: center;
    margin-bottom: 18px;
    transition: all 0.3s ease;
}
.metric-card:hover .metric-icon {
    background: rgba(197,165,102,0.13);
    border-color: rgba(197,165,102,0.4);
}
.metric-icon svg {
    width: 18px; height: 18px;
    stroke: var(--g); fill: none; stroke-width: 1.5;
    stroke-linecap: round; stroke-linejoin: round;
}
.metric-card h4 {
    font-family: var(--sans);
    font-size: 0.72rem; font-weight: 600;
    letter-spacing: 3.5px; text-transform: uppercase;
    color: rgba(197,165,102,0.45); margin: 0 0 8px 0;
}
.metric-card h2 {
    font-family: var(--serif);
    font-size: 3rem; font-weight: 600;
    line-height: 1; margin: 0;
    background: linear-gradient(135deg, var(--g2) 0%, var(--g) 100%);
    background-clip: text; -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
}
.metric-card-sub {
    font-size: 0.78rem; color: rgba(197,165,102,0.3);
    letter-spacing: 0.8px; margin-top: 6px;
    font-style: italic;
}

/* ═══ DOWNLOAD SECTION ═════════════════════════════════════════════════════ */
.dl-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }

[data-testid="stDownloadButton"] > button {
    background: var(--p3) !important;
    border: 1px solid rgba(197,165,102,0.14) !important;
    color: rgba(197,165,102,0.8) !important;
    border-radius: var(--r-md) !important;
    font-family: var(--sans) !important;
    font-size: 0.78rem !important; font-weight: 500 !important;
    letter-spacing: 1px !important; text-transform: uppercase !important;
    padding: 14px 18px !important;
    transition: all 0.3s cubic-bezier(.22,1,.36,1) !important;
    width: 100% !important; text-align: left !important;
    position: relative !important; overflow: hidden !important;
}
/* Linha lateral de acento */
[data-testid="stDownloadButton"] > button::before {
    content: '';
    position: absolute; top: 12%; left: 0;
    width: 3px; height: 76%;
    background: linear-gradient(180deg, transparent, var(--g), transparent);
    border-radius: 0 2px 2px 0; opacity: 0.4;
    transition: opacity 0.25s ease;
}
/* Reflexo de entrada */
[data-testid="stDownloadButton"] > button::after {
    content: '';
    position: absolute; top: 0; left: -100%;
    width: 60%; height: 100%;
    background: linear-gradient(90deg, transparent, rgba(197,165,102,0.06), transparent);
    transition: left 0.45s ease;
    transform: skewX(-15deg);
}
[data-testid="stDownloadButton"] > button:hover {
    background: var(--p4) !important;
    border-color: rgba(197,165,102,0.35) !important;
    color: var(--g) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 28px rgba(0,0,0,0.35), 0 0 16px rgba(197,165,102,0.06) !important;
}
[data-testid="stDownloadButton"] > button:hover::before { opacity: 1; }
[data-testid="stDownloadButton"] > button:hover::after  { left: 100%; }

/* ═══ DATAFRAME ════════════════════════════════════════════════════════════ */
[data-testid="stDataFrame"] {
    border: 1px solid rgba(197,165,102,0.12) !important;
    border-radius: var(--r-md) !important;
    overflow: hidden !important;
    animation: fadeUp 0.4s ease both;
}
[data-testid="stDataFrame"] th {
    background: var(--p3) !important;
    color: rgba(197,165,102,0.65) !important;
    font-family: var(--sans) !important;
    font-size: 0.72rem !important; font-weight: 600 !important;
    letter-spacing: 2px !important; text-transform: uppercase !important;
    border-bottom: 1px solid rgba(197,165,102,0.15) !important;
    padding: 14px 16px !important;
}
[data-testid="stDataFrame"] td {
    font-family: var(--sans) !important;
    font-size: 0.86rem !important;
    color: rgba(237,229,212,0.75) !important;
    border-color: rgba(197,165,102,0.06) !important;
    padding: 11px 16px !important;
}
[data-testid="stDataFrame"] tr:hover td {
    background: rgba(197,165,102,0.04) !important;
}

/* ═══ ALERT / INFO ═════════════════════════════════════════════════════════ */
[data-testid="stAlert"] {
    background: rgba(197,165,102,0.05) !important;
    border: 1px solid rgba(197,165,102,0.18) !important;
    border-radius: var(--r-md) !important;
    color: rgba(197,165,102,0.7) !important;
    font-size: 0.82rem !important;
}

/* ═══ SPINNER ══════════════════════════════════════════════════════════════ */
[data-testid="stSpinner"] > div {
    border-color: rgba(197,165,102,0.15) !important;
    border-top-color: var(--g) !important;
    border-radius: 50%;
}
[data-testid="stSpinner"] p {
    color: rgba(197,165,102,0.5) !important;
    font-family: var(--sans) !important;
    font-size: 0.82rem !important; letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
}

/* ═══ SIDEBAR ══════════════════════════════════════════════════════════════ */
[data-testid="stSidebar"] {
    background: #04060C;
    border-right: 1px solid rgba(197,165,102,0.1);
}
[data-testid="stSidebar"] > div:first-child { padding-top: 0; }

.sb-header {
    padding: 28px 18px 18px;
    border-bottom: 1px solid rgba(197,165,102,0.08);
    position: relative;
    background: linear-gradient(180deg,
        rgba(197,165,102,0.04) 0%, transparent 100%);
}
.sb-header::after {
    content: '';
    position: absolute; bottom: -1px; left: 0;
    width: 50px; height: 1px;
    background: var(--g);
    animation: glow 2.5s ease-in-out infinite;
}
.sb-eyebrow {
    font-size: 0.65rem; font-weight: 600;
    letter-spacing: 4px; text-transform: uppercase;
    color: rgba(197,165,102,0.32); margin-bottom: 4px;
}
.sb-title {
    font-family: var(--serif);
    font-size: 1.3rem; font-weight: 600;
    color: var(--c); letter-spacing: 0.5px; margin: 0;
}
.sb-title span { color: var(--g); }

/* Tutorial na sidebar */
.sb-tutorial {
    margin: 12px 14px;
    padding: 12px 14px;
    border-radius: var(--r-sm);
    background: rgba(197,165,102,0.04);
    border: 1px solid rgba(197,165,102,0.1);
}
.sb-tutorial-title {
    font-size: 0.68rem; font-weight: 600;
    letter-spacing: 2.5px; text-transform: uppercase;
    color: rgba(197,165,102,0.4); margin-bottom: 7px;
}
.sb-tutorial-text {
    font-size: 0.82rem; color: rgba(237,229,212,0.38);
    line-height: 1.6; letter-spacing: 0.2px;
}

/* Botões marcar/desmarcar */
[data-testid="stSidebar"] .stButton > button {
    border-radius: var(--r-sm) !important;
    font-family: var(--sans) !important;
    font-size: 0.57rem !important; font-weight: 600 !important;
    letter-spacing: 1.8px !important; text-transform: uppercase !important;
    padding: 7px 0 !important; width: 100% !important;
    transition: all 0.2s ease !important; border: none !important;
}
[data-testid="stSidebar"] .stButton:nth-of-type(1) > button {
    background: rgba(197,165,102,0.1) !important;
    color: var(--g) !important;
    border-right: 1px solid rgba(197,165,102,0.12) !important;
}
[data-testid="stSidebar"] .stButton:nth-of-type(1) > button:hover {
    background: rgba(197,165,102,0.18) !important;
    transform: none !important;
}
[data-testid="stSidebar"] .stButton:nth-of-type(2) > button {
    background: transparent !important;
    color: rgba(237,229,212,0.22) !important;
}
[data-testid="stSidebar"] .stButton:nth-of-type(2) > button:hover {
    color: rgba(237,229,212,0.5) !important;
    background: rgba(255,255,255,0.025) !important;
}

/* Checkboxes como items de lista */
[data-testid="stSidebar"] .stCheckbox { margin:0 !important; padding:0 !important; }
[data-testid="stSidebar"] .stCheckbox > label > div:first-child { display:none !important; }
[data-testid="stSidebar"] .stCheckbox > label {
    display: flex !important; align-items: center !important;
    gap: 8px !important; width: 100% !important;
    padding: 8px 14px 8px 16px !important;
    margin: 0 !important; cursor: pointer !important;
    border-left: 2px solid transparent !important;
    border-radius: 0 var(--r-sm) var(--r-sm) 0 !important;
    transition: all 0.18s ease !important;
    background: transparent !important;
}
[data-testid="stSidebar"] .stCheckbox > label:hover {
    background: rgba(197,165,102,0.04) !important;
    border-left-color: rgba(197,165,102,0.22) !important;
}
[data-testid="stSidebar"] .stCheckbox > label > div:last-child,
[data-testid="stSidebar"] .stCheckbox > label > span {
    font-family: var(--sans) !important;
    font-size: 0.8rem !important; font-weight: 400 !important;
    color: rgba(237,229,212,0.28) !important;
    letter-spacing: 0.8px !important; text-transform: uppercase !important;
    line-height: 1 !important;
}
[data-testid="stSidebar"] .stCheckbox:has(input:checked) > label {
    border-left-color: var(--g) !important;
    background: rgba(197,165,102,0.06) !important;
}
[data-testid="stSidebar"] .stCheckbox:has(input:checked) > label > div:last-child,
[data-testid="stSidebar"] .stCheckbox:has(input:checked) > label > span {
    color: rgba(197,165,102,0.85) !important; font-weight: 500 !important;
}

.sidebar-divider {
    border: none;
    border-top: 1px solid rgba(197,165,102,0.07);
    margin: 4px 0 0;
}
.rubrica-count {
    padding: 8px 16px 16px;
    font-size: 0.68rem; font-family: var(--sans);
    letter-spacing: 2px; text-transform: uppercase;
}

/* ═══ FOOTER ═══════════════════════════════════════════════════════════════ */
.em-footer {
    margin-top: 100px;
    padding: 48px 0 24px;
    border-top: 1px solid rgba(197,165,102,0.1);
    display: flex; flex-direction: column;
    align-items: center; gap: 16px;
    position: relative;
}
.em-footer::before {
    content: '';
    position: absolute; top: -1px; left: 50%;
    transform: translateX(-50%);
    width: 100px; height: 1px;
    background: linear-gradient(90deg, transparent, var(--g), transparent);
    animation: glow 3s ease-in-out infinite;
}
.em-footer-name {
    font-family: var(--serif);
    font-size: 1.8rem; font-weight: 600; font-style: italic;
    color: var(--g); letter-spacing: 2px;
}
.em-footer-contacts {
    display: flex; gap: 28px; flex-wrap: wrap; justify-content: center;
}
.em-footer-contact {
    font-size: 0.8rem; color: rgba(197,165,102,0.4); letter-spacing: 1px;
}
.em-footer-contact a {
    color: rgba(197,165,102,0.4); text-decoration: none;
    transition: color 0.2s;
}
.em-footer-contact a:hover { color: var(--g); }
.em-whatsapp-btn {
    display: inline-flex; align-items: center; gap: 10px;
    background: rgba(197,165,102,0.05);
    border: 1px solid rgba(197,165,102,0.3);
    border-radius: var(--r-sm);
    color: rgba(197,165,102,0.7);
    font-family: var(--sans); font-size: 0.64rem;
    font-weight: 600; letter-spacing: 2.5px;
    text-transform: uppercase; padding: 11px 28px;
    text-decoration: none;
    transition: all 0.3s cubic-bezier(.22,1,.36,1);
    cursor: pointer; margin-top: 6px; position: relative; overflow: hidden;
}
.em-whatsapp-btn::after {
    content: '';
    position: absolute; top: 0; left: -100%;
    width: 60%; height: 100%;
    background: linear-gradient(90deg, transparent, rgba(197,165,102,0.08), transparent);
    transform: skewX(-15deg);
    transition: left 0.5s ease;
}
.em-whatsapp-btn:hover {
    border-color: var(--g); color: var(--g); text-decoration: none;
    background: rgba(197,165,102,0.1);
    box-shadow: 0 8px 30px rgba(197,165,102,0.1);
    transform: translateY(-2px);
}
.em-whatsapp-btn:hover::after { left: 120%; }

/* ═══ SCROLLBAR ════════════════════════════════════════════════════════════ */
::-webkit-scrollbar { width: 3px; height: 3px; }
::-webkit-scrollbar-track { background: var(--p); }
::-webkit-scrollbar-thumb {
    background: rgba(197,165,102,0.22);
    border-radius: 2px;
}
::-webkit-scrollbar-thumb:hover { background: rgba(197,165,102,0.5); }

/* ═══ HEADER NATIVO ════════════════════════════════════════════════════════ */
header[data-testid="stHeader"] {
    background: rgba(4,6,12,0.9) !important;
    border-bottom: 1px solid rgba(197,165,102,0.07) !important;
    backdrop-filter: blur(16px) !important;
}

/* ═══ SELO DE FUNDAÇÃO — fixo, canto inferior direito ══════════════════════ */
.em-founder-seal {
    position: fixed;
    bottom: 20px;
    right: 24px;
    z-index: 9999;
    display: flex;
    flex-direction: column;
    align-items: flex-end;
    gap: 2px;
    pointer-events: none;
    opacity: 0.65;
    transition: opacity 0.4s ease;
}
.em-founder-seal:hover {
    opacity: 1;
    pointer-events: auto;
}
.em-seal-line {
    width: 100%;
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(197,165,102,0.5));
    margin-bottom: 6px;
    animation: glow 3s ease-in-out infinite;
}
.em-seal-label {
    font-family: 'Inter', sans-serif;
    font-size: 0.46rem;
    font-weight: 600;
    letter-spacing: 3.5px;
    text-transform: uppercase;
    color: rgba(197,165,102,0.35);
    text-align: right;
}
.em-seal-name {
    font-family: 'Great Vibes', cursive;
    font-size: 1.6rem;
    color: rgba(197,165,102,0.6);
    line-height: 1;
    text-align: right;
}
.em-seal-sub {
    font-family: 'Inter', sans-serif;
    font-size: 0.42rem;
    font-weight: 500;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: rgba(197,165,102,0.22);
    text-align: right;
    margin-top: 1px;
}
.em-seal-ornament {
    font-size: 0.35rem;
    color: rgba(197,165,102,0.28);
    letter-spacing: 4px;
    margin-top: 4px;
}
</style>
""", unsafe_allow_html=True)

# --- 1b. LÓGICA DE LOGIN ---
def _check_login(email: str, senha: str) -> bool:
    return email.strip() == "edson.senabr@gmail.com" and senha == "Edsonsena14"

if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False

if not st.session_state["autenticado"]:

    # CSS da tela de login — Painel Completo v1
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,400;0,500;0,600;0,700;1,400;1,600&family=Inter:wght@300;400;500;600;700&family=Great+Vibes&display=swap');

    header,footer,[data-testid="stSidebar"],[data-testid="stToolbar"],
    [data-testid="stDecoration"],[data-testid="stStatusWidget"],
    [data-testid="stHeader"]{display:none!important;}

    /* Fundo */
    .stApp{
        background:#030508!important;
        background-image:
            linear-gradient(rgba(197,165,102,.016) 1px,transparent 1px),
            linear-gradient(90deg,rgba(197,165,102,.016) 1px,transparent 1px)!important;
        background-size:56px 56px!important;
    }

    /* Reset de padding — todos os seletores do Streamlit */
    .block-container,.main .block-container,div.block-container{
        padding:0!important;margin:0!important;
        max-width:100%!important;width:100%!important;min-height:0!important;
    }
    section[data-testid="stMain"],
    section[data-testid="stMain"]>div,
    section[data-testid="stMain"]>div>div,
    [data-testid="stAppViewBlockContainer"],
    [data-testid="stMainBlockContainer"],
    .appview-container .main,
    .appview-container section{
        padding:0!important;margin:0!important;
        max-width:100%!important;width:100%!important;
    }
    [data-testid="stHorizontalBlock"]{
        gap:0!important;padding:0!important;margin:0!important;
        align-items:stretch!important;
    }
    [data-testid="stHorizontalBlock"]>div{padding:0!important;margin:0!important;}
    [data-testid="stVerticalBlock"]{gap:0!important;padding:0!important;margin:0!important;}
    [data-testid="stVerticalBlock"]>div,
    [data-testid="stVerticalBlock"]>div>div{padding:0!important;margin:0!important;}
    [data-testid="stMarkdown"]{margin:0!important;padding:0!important;line-height:normal!important;}
    [data-testid="stMarkdown"]>div{margin:0!important;padding:0!important;}
    div.stMarkdown,div.stMarkdown>div{margin:0!important;padding:0!important;}
    [data-testid="stForm"]{background:transparent!important;border:none!important;padding:0!important;margin:0!important;}
    [data-testid="stForm"]>div,[data-testid="stForm"]>div>div{padding:0!important;margin:0!important;}
    div[data-testid="stTextInput"]{margin-bottom:8px!important;}
    div[data-testid="stTextInput"]>div{padding:0!important;}

    /* ══ ANIMAÇÕES ══ */
    @keyframes pulseDot{0%,100%{opacity:.25;transform:scale(1)}50%{opacity:1;transform:scale(1.3)}}
    @keyframes shimmer{0%{background-position:-500px 0}100%{background-position:500px 0}}
    @keyframes fadeUp{from{opacity:0;transform:translateY(14px)}to{opacity:1;transform:translateY(0)}}
    @keyframes glowLine{0%,100%{opacity:.3}50%{opacity:.8}}

    /* ══ COLUNA ESQUERDA — Produto ══ */
    .lp-left{
        background:linear-gradient(150deg,rgba(197,165,102,.065) 0%,rgba(3,5,8,.02) 50%,rgba(3,5,8,0) 100%);
        border-right:1px solid rgba(197,165,102,.1);
        padding:48px 36px 40px 44px;
        min-height:100vh;
        display:flex;flex-direction:column;
        justify-content:space-between;
        position:relative;overflow:hidden;
    }
    .lp-left::after{
        content:'';position:absolute;
        bottom:-100px;right:-100px;
        width:320px;height:320px;border-radius:50%;
        border:1px solid rgba(197,165,102,.06);pointer-events:none;
    }

    /* Marca */
    .lp-brand{display:flex;align-items:center;gap:11px;margin-bottom:32px;animation:fadeUp .6s ease both;}
    .lp-mark{
        width:36px;height:36px;background:rgba(197,165,102,.1);
        border:1px solid rgba(197,165,102,.24);border-radius:8px;
        display:flex;align-items:center;justify-content:center;
        font-family:'Cormorant Garamond',serif;font-size:14px;font-weight:700;color:#C5A566;
    }
    .lp-bname{font-family:'Inter',sans-serif;font-size:9px;letter-spacing:3px;text-transform:uppercase;color:rgba(237,229,212,.25);font-weight:600;}

    /* Nome e tagline do produto */
    .lp-pname{
        font-family:'Cormorant Garamond',serif;
        font-size:54px;font-weight:600;line-height:.92;
        color:#EDE5D4;margin-bottom:6px;
        animation:fadeUp .6s .1s ease both;
    }
    .lp-pname span{
        background:linear-gradient(110deg,#EDE5D4 10%,#C5A566 35%,#D4B87A 55%,#EDE5D4 75%);
        background-size:500px;
        -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;
        animation:shimmer 5s linear infinite;
    }
    .lp-ptag{
        font-family:'Cormorant Garamond',serif;font-style:italic;
        font-size:15px;color:rgba(197,165,102,.42);
        margin-bottom:24px;letter-spacing:.5px;
        animation:fadeUp .6s .15s ease both;
    }

    /* Badge de velocidade */
    .lp-speed{
        display:flex;align-items:center;gap:12px;
        padding:12px 16px;
        background:rgba(197,165,102,.05);
        border:1px solid rgba(197,165,102,.15);
        border-left:3px solid rgba(197,165,102,.6);
        margin-bottom:24px;
        animation:fadeUp .6s .2s ease both;
    }
    .lp-speed-n{
        font-family:'Cormorant Garamond',serif;
        font-size:36px;font-weight:600;color:#C5A566;line-height:1;flex-shrink:0;
    }
    .lp-speed-t{font-family:'Inter',sans-serif;font-size:12px;font-weight:600;color:rgba(237,229,212,.55);}
    .lp-speed-s{font-family:'Inter',sans-serif;font-size:10px;color:rgba(237,229,212,.25);margin-top:3px;line-height:1.5;}

    /* Vantagens numeradas */
    .lp-bens{display:flex;flex-direction:column;gap:0;animation:fadeUp .6s .25s ease both;}
    .lp-ben{
        display:flex;align-items:flex-start;gap:12px;
        padding:12px 0;
        border-bottom:1px solid rgba(197,165,102,.07);
    }
    .lp-ben:last-child{border-bottom:none;}
    .lp-bnum{
        font-family:'Cormorant Garamond',serif;
        font-size:18px;font-weight:300;
        color:rgba(197,165,102,.22);min-width:22px;line-height:1.3;flex-shrink:0;
    }
    .lp-btitle{
        font-family:'Inter',sans-serif;font-size:12px;font-weight:600;
        color:rgba(237,229,212,.55);margin-bottom:3px;letter-spacing:.2px;
    }
    .lp-bdesc{font-family:'Inter',sans-serif;font-size:11px;color:rgba(237,229,212,.24);line-height:1.55;}

    /* Rodapé esquerdo */
    .lp-lfoot{
        font-family:'Inter',sans-serif;font-size:9px;
        color:rgba(197,165,102,.18);letter-spacing:.8px;line-height:1.9;
        margin-top:auto;padding-top:20px;
        animation:fadeUp .6s .3s ease both;
    }

    /* ══ COLUNA DIREITA — Login ══ */
    .lp-right{
        background:rgba(4,6,12,.97);
        padding:48px 44px 40px 40px;
        min-height:100vh;
        display:flex;flex-direction:column;
        justify-content:space-between;
    }

    /* Logo */
    .lp-r-logo{
        font-family:'Cormorant Garamond',serif;
        font-size:38px;font-weight:600;color:#EDE5D4;line-height:1;
        margin-bottom:4px;animation:fadeUp .6s .1s ease both;
    }
    .lp-r-logo span{color:#C5A566;}
    .lp-r-sub{
        font-family:'Cormorant Garamond',serif;font-style:italic;
        font-size:13px;color:rgba(197,165,102,.32);
        margin-bottom:28px;letter-spacing:.8px;
        animation:fadeUp .6s .15s ease both;
    }
    .lp-r-orn{
        display:flex;align-items:center;gap:10px;margin-bottom:24px;
        animation:fadeUp .6s .2s ease both;
    }
    .lp-r-ol{flex:1;height:1px;background:rgba(197,165,102,.1);}
    .lp-r-od{font-size:7px;color:rgba(197,165,102,.24);}

    /* Inputs Streamlit */
    [data-testid="stForm"] label{
        font-family:'Inter',sans-serif!important;font-size:9px!important;
        font-weight:600!important;letter-spacing:3px!important;text-transform:uppercase!important;
        color:rgba(197,165,102,.36)!important;
    }
    [data-testid="stForm"] label span{display:none!important;}
    [data-testid="stForm"] input{
        background:rgba(197,165,102,.04)!important;
        border:1px solid rgba(197,165,102,.15)!important;
        border-radius:8px!important;color:#EDE5D4!important;
        font-family:'Inter',sans-serif!important;font-size:14px!important;font-weight:300!important;
        padding:12px 16px!important;caret-color:#C5A566!important;
        outline:none!important;box-shadow:none!important;transition:all .25s!important;
    }
    [data-testid="stForm"] input:focus{
        border-color:rgba(197,165,102,.5)!important;
        background:rgba(197,165,102,.07)!important;
        box-shadow:0 0 0 3px rgba(197,165,102,.06)!important;
    }
    [data-testid="stFormSubmitButton"]>button{
        width:100%!important;background:rgba(197,165,102,.11)!important;
        border:1px solid rgba(197,165,102,.4)!important;border-radius:8px!important;
        color:#C5A566!important;font-family:'Inter',sans-serif!important;
        font-size:10px!important;font-weight:700!important;
        letter-spacing:3.5px!important;text-transform:uppercase!important;
        padding:14px!important;margin-top:12px!important;transition:all .3s!important;
    }
    [data-testid="stFormSubmitButton"]>button:hover{
        background:rgba(197,165,102,.2)!important;border-color:#C5A566!important;
        transform:translateY(-2px)!important;
        box-shadow:0 10px 28px rgba(197,165,102,.14)!important;
    }
    [data-testid="stAlert"]{
        background:rgba(180,60,60,.05)!important;border:1px solid rgba(180,60,60,.2)!important;
        border-radius:8px!important;color:rgba(220,110,110,.8)!important;font-size:11px!important;
    }
    [data-testid="stAlert"] svg{display:none!important;}

    /* Grid de stats */
    .lp-r-grid{
        display:grid;grid-template-columns:repeat(3,1fr);gap:1px;
        background:rgba(197,165,102,.09);border-radius:10px;overflow:hidden;
        margin-top:20px;animation:fadeUp .6s .35s ease both;
    }
    .lp-gs{background:#04060C;padding:14px 14px;}
    .lp-gsn{
        font-family:'Cormorant Garamond',serif;
        font-size:22px;font-weight:600;color:#C5A566;line-height:1;
    }
    .lp-gsl{
        font-family:'Inter',sans-serif;font-size:8px;font-weight:600;
        letter-spacing:2px;text-transform:uppercase;
        color:rgba(237,229,212,.18);margin-top:4px;
    }

    /* Por que usar — rodapé direito */
    .lp-r-why{
        margin-top:16px;padding-top:14px;
        border-top:1px solid rgba(197,165,102,.07);
        animation:fadeUp .6s .4s ease both;
    }
    .lp-r-why-label{
        font-family:'Inter',sans-serif;font-size:8px;font-weight:600;
        letter-spacing:3px;text-transform:uppercase;
        color:rgba(197,165,102,.22);margin-bottom:10px;
    }
    .lp-r-why-row{display:flex;gap:0;}
    .lp-r-wi{
        flex:1;padding:0 10px;
        border-right:1px solid rgba(197,165,102,.07);text-align:center;
    }
    .lp-r-wi:first-child{padding-left:0;text-align:left;}
    .lp-r-wi:last-child{border-right:none;}
    .lp-r-win{
        font-family:'Cormorant Garamond',serif;
        font-size:16px;font-weight:600;color:#C5A566;line-height:1;
    }
    .lp-r-wil{
        font-family:'Inter',sans-serif;font-size:8px;
        color:rgba(237,229,212,.18);margin-top:3px;line-height:1.4;
    }

    /* Status */
    .lp-r-status{
        display:flex;align-items:center;gap:6px;margin-top:14px;
        font-family:'Inter',sans-serif;font-size:8px;
        color:rgba(197,165,102,.18);letter-spacing:1.5px;text-transform:uppercase;
    }
    .lp-r-sdot{
        width:5px;height:5px;border-radius:50%;background:#4CAF50;
        animation:pulseDot 2.5s ease-in-out infinite;
    }
    @keyframes pulseDot{0%,100%{opacity:.3;transform:scale(1)}50%{opacity:1;transform:scale(1.2)}}

    /* Selo */
    .seal{position:fixed;bottom:40px;right:24px;z-index:9999;display:flex;flex-direction:column;align-items:flex-end;gap:2px;pointer-events:none;opacity:.55;transition:opacity .4s;}
    .seal:hover{opacity:1;pointer-events:auto;}
    .seal-line{width:100%;height:1px;background:linear-gradient(90deg,transparent,rgba(197,165,102,.5));margin-bottom:4px;}
    .seal-label{font-family:'Inter',sans-serif;font-size:7px;font-weight:600;letter-spacing:3px;text-transform:uppercase;color:rgba(197,165,102,.28);text-align:right;}
    .seal-name{font-family:'Great Vibes',cursive;font-size:20px;color:rgba(197,165,102,.52);line-height:1;text-align:right;}
    .seal-sub{font-family:'Inter',sans-serif;font-size:6px;letter-spacing:2px;text-transform:uppercase;color:rgba(197,165,102,.18);text-align:right;margin-top:1px;}
    </style>
    """, unsafe_allow_html=True)

    # ══ LAYOUT: esquerda (produto) | direita (login) ══
    col_l, col_r = st.columns([1.15, 1.0])

    # ════════ COLUNA ESQUERDA — PRODUTO ════════
    with col_l:
        st.markdown('<div class="lp-left">', unsafe_allow_html=True)

        # Marca
        st.markdown('<div class="lp-brand"><div class="lp-mark">EM</div><div class="lp-bname">Edson Medeiros · Consultorias</div></div>', unsafe_allow_html=True)

        # Nome + tagline
        st.markdown('<div class="lp-pname">Extrato<span>X</span></div>', unsafe_allow_html=True)
        st.markdown('<div class="lp-ptag">O robô que audita seu extrato em 10 segundos.</div>', unsafe_allow_html=True)

        # Badge de velocidade
        st.markdown('<div class="lp-speed"><div class="lp-speed-n">10s</div><div><div class="lp-speed-t">Análise completa do extrato</div><div class="lp-speed-s">Do upload ao relatório pronto —<br>sem nenhum esforço manual</div></div></div>', unsafe_allow_html=True)

        # Vantagens numeradas — cada uma separada
        st.markdown('<div class="lp-bens">', unsafe_allow_html=True)
        vantagens = [
            ("01", "Precisão de 100%",          "Leitura posicional por coluna X — distingue débito de crédito sem falhas e sem interpretação errada"),
            ("02", "19 rubricas monitoradas",    "CESTA, MORA, ANUIDADE, ENCARGOS, PARCELA, SEGURO, SAQUE e mais — todas detectadas automaticamente"),
            ("03", "Relatório jurídico pronto",  "Planilha com valor em dobro (Art. 42 CDC) gerada instantaneamente, pronta para peticionamento"),
        ]
        for num, titulo, desc in vantagens:
            st.markdown(f'<div class="lp-ben"><div class="lp-bnum">{num}</div><div><div class="lp-btitle">{titulo}</div><div class="lp-bdesc">{desc}</div></div></div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        # Rodapé
        st.markdown('<div class="lp-lfoot">(92) 99508-7379 &nbsp;·&nbsp; edson.senabr@gmail.com</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # ════════ COLUNA DIREITA — LOGIN ════════
    with col_r:
        st.markdown('<div class="lp-right">', unsafe_allow_html=True)

        # Logo
        st.markdown('<div class="lp-r-logo">Extrato<span>X</span></div>', unsafe_allow_html=True)
        st.markdown('<div class="lp-r-sub">Acesse o portal de auditoria</div>', unsafe_allow_html=True)
        st.markdown('<div class="lp-r-orn"><div class="lp-r-ol"></div><div class="lp-r-od">◆</div><div class="lp-r-ol"></div></div>', unsafe_allow_html=True)

        # Formulário
        with st.form("login_form", clear_on_submit=False):
            _email = st.text_input("E-mail", placeholder="seu@email.com", key="login_email")
            _senha = st.text_input("Senha", placeholder="••••••••••", type="password", key="login_senha")
            _submitted = st.form_submit_button("◆  Acessar o ExtratoX")

        if _submitted:
            if _check_login(_email, _senha):
                st.session_state["autenticado"] = True
                st.rerun()
            else:
                st.error("Credenciais inválidas — verifique e-mail e senha")

        # Grid de stats
        st.markdown('<div class="lp-r-grid"><div class="lp-gs"><div class="lp-gsn">10s</div><div class="lp-gsl">Análise</div></div><div class="lp-gs"><div class="lp-gsn">19</div><div class="lp-gsl">Rubricas</div></div><div class="lp-gs"><div class="lp-gsn">Art.42</div><div class="lp-gsl">CDC auto</div></div></div>', unsafe_allow_html=True)

        # Por que usar
        st.markdown('<div class="lp-r-why"><div class="lp-r-why-label">Por que usar</div><div class="lp-r-why-row"><div class="lp-r-wi"><div class="lp-r-win">100%</div><div class="lp-r-wil">Precisão</div></div><div class="lp-r-wi"><div class="lp-r-win">0</div><div class="lp-r-wil">Esforço manual</div></div><div class="lp-r-wi"><div class="lp-r-win">CDC</div><div class="lp-r-wil">Art. 42 auto</div></div></div></div>', unsafe_allow_html=True)

        # Status
        st.markdown('<div class="lp-r-status"><div class="lp-r-sdot"></div>Sistema online &nbsp;·&nbsp; Versão 2.0</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Selo
    st.markdown('<div class="seal"><div class="seal-line"></div><div class="seal-label">Fundado por</div><div class="seal-name">Edson Medeiros</div><div class="seal-sub">Consultorias &amp; Compliance · 2024</div></div>', unsafe_allow_html=True)

    st.stop()

# --- (USUÁRIO AUTENTICADO — APP NORMAL A PARTIR DAQUI) ---

# --- 2. RÚBRICAS ---
RUBRICAS_MESTRE = {
    # "TARIFA BANCARIA / CESTA B.EXPRESSO4" — o nome real no extrato está na sublinha
    "CESTA": r"\bCESTA\b",

    # "PACOTE DE SERVICOS" ou "PACOTE SERVICOS"
    "PACOTE": r"\bPACOTE\b",

    # "MORA DE OPERACAO" / "MORA OPERACAO"
    "MORA DE OPERAÇÃO": r"MORA\s+DE\s+OPERA[CÇ]AO|MORA\s+OPERA[CÇ]AO\b",

    # "MORA CREDITO PESSOAL" / "MORA CRED PESS" / "MORA CP"
    "MORA CREDITO PESSOAL": r"MORA\s+CREDITO\s+PESSOAL|MORA\s+CRED\s+PESS|MORA\s+CP\b",

    # "MORA OPERACAO DE CREDITO" / "MORA OPER CRED"
    "MORA OPERACAO DE CREDITO": r"MORA\s+OPERA[CÇ]AO\s+DE\s+CREDITO|MORA\s+OPER\s+CRED",

    # "BX" isolado — word boundary para não pegar "BXA" ou "COBRA"
    "BX": r"\bBX\b",

    # "PARCELA CREDITO PESSOAL" / "PARC CRED PESS" / "PARCELA CP"
    "PARCELA CREDITO PESSOAL": r"PARCELA\s+CREDITO\s+PESSOAL|PARC\s+CRED\s+PESS|PARCELA\s+CP\b",

    # "GASTOS CARTAO DE CREDITO" / "CARTAO DE CREDITO" / "FATURA CARTAO"
    # NÃO inclui "CARTAO CREDITO ANUIDADE" (já capturado em ANUIDADE)
    "GASTOS CARTAO DE CREDITO": r"GASTOS\s+CART[AÃ]O|FATURA\s+CART[AÃ]O|CART[AÃ]O\s+DE\s+CREDITO(?!\s+ANUIDADE)",

    # "SEGURO" / "SEG " / "SEGURADORA" — word boundary para não pegar "SAQUE"
    "SEGURO": r"\bSEGURO\b|\bSEGURADORA\b|\bSEG\s",

    # "ADIANT" / "ADIANTAMENTO"
    "ADIANT": r"\bADIANT|\bADIANTAMENTO\b",

    # "APLICACAO" / "APLIC" isolado
    "APLIC": r"\bAPLICA[CÇ]AO\b|\bAPLIC\b",

    # "ENCARGOS" / "ENCARGO" / "ENCARGOS LIMITE DE CRED" / "IOF" não — só encargo mesmo
    "ENCARGOS": r"\bENCARGOS?\b|\bENC\s+LIMITE\b|\bLIMITE\s+DE\s+CRED\b",

    # "CARTAO CREDITO ANUIDADE" / "ANUIDADE" — verificado ANTES de GASTOS CARTAO
    "ANUIDADE": r"\bANUIDADE\b|CART[AÃ]O\s+CREDITO\s+ANUIDADE",

    # "OPERACOES VENCIDAS" / "OPERAÇÕES VENCIDAS"
    "OPERACOES VENCIDAS": r"OPERA[CÇ][OÕ]ES\s+VENCIDAS",

    # "BRADESCO VIDA E PREVIDENCIA" / "VIDA E PREVIDENCIA" / "APORTE VGBL" / "PAGTO BRADESCO VIDA"
    "BRADESCO VIDA E PREVIDENCIA": r"BRADESCO\s+VIDA|VIDA\s+E\s+PREVID[EÊ]NCIA|APORTE\s+VGBL|PAGTO.*VIDA",

    # "TITULO DE CAPITALIZACAO" / variações com acento
    "TITULO DE CAPITALIZACAO": r"T[IÍ]TULO\s+DE\s+CAPITALIZ|\bCAPITALIZ[AÇ]",

    # "AUTO RE" — seguro automóvel / renovação automática
    "AUTO RE": r"\bAUTO\s+RE\b|\bAUTORE\b",

    # "SAQUEcorrespondente" / "SAQUEterminal" / "SAQUEpremium" — tarifa de saque bancário
    # Aparecem como SUBLINHA de "TARIFA BANCARIA" no extrato Bradesco (Caso C):
    #   Linha anterior (COM valor): "TARIFA BANCARIA  0000001  6,75  2.239,89"
    #   Sublinha (sem valor):       "SAQUEcorrespondente"   ← esta linha é detectada
    #
    # O motor Caso C: rubrica sem valor na linha atual → busca valor na linha ANTERIOR.
    # A linha anterior "TARIFA BANCARIA X,XX" fornece o valor correto.
    #
    # NÃO captura "SAQUE DIN CORBAN CARTAO" (saque do cliente — não é tarifa).
    # NÃO captura "SAQUE DINHEIRO ATM" (saque do cliente — não é tarifa).
    # NÃO captura "CESTA B.EXPRESSO4" (sublinha diferente — capturada pela rubrica CESTA).
    "SAQUE TERMINAL": r"\bSAQUECORRESPONDENTE\b|\bSAQUETERMINAL\b|\bSAQUE\s+CORRESPONDENTE\b|\bSAQUE\s+TERMINAL\b|\bSAQUEPREMIUM\b",

    # "TARIFA EMISSAO EXTRATO" / "EXTRATOmes(E)" — tarifa de emissão de extrato mensal
    # No extrato Bradesco aparece em duas formas:
    #   Forma A (linha completa): "TARIFA EMISSAO EXTRATO 0210220 1,35 1,00" → TEM o valor
    #   Forma B (sublinha):       "EXTRATOMES(E)"  →  sem valor, a linha anterior TEM o valor
    #
    # ATENÇÃO: capturamos APENAS a Forma A (linha completa com valor).
    # A sublinha EXTRATOMES(E) é ignorada pois o valor já foi capturado na linha anterior.
    # Isso evita duplicatas e valores errados causados pelo Caso C pegando lançamentos vizinhos.
    "EXTRATO MES": r"TARIFA\s+EMISSAO\s+EXTRATO",
}

TERMOS_EXCLUSAO = r"TRANSF|SALDO|SDO|TRANSFERENCIA|SALARIO"

# --- 3. MOTOR — LÓGICA DATA INFERIOR ---
#
# COMO FUNCIONA O MODELO "DATA INFERIOR":
#
# No extrato Bradesco, existem dois formatos de linha:
#
#   FORMATO A — linha COM data ao lado da rubrica:
#     "15/01/2020  TARIFA BANCARIA  CESTA B.EXPRESSO4  21,60"
#     → rubrica e data estão juntas. A data pertence a esse lançamento.
#
#   FORMATO B — linha SEM data (rubrica "solta"):
#     "MORA CREDITO PESSOAL  115,62"
#     "ENCARGOS LIMITE DE CRED  19,31"
#     "08/02/2017  SAQUE DIN CORBAN CARTAO  ..."   ← próxima linha datada
#
#   No formato B, as rubricas acima não têm data própria.
#   A data que as referencia é a da PRÓXIMA linha que contiver uma data —
#   chamada aqui de "data inferior" pois aparece abaixo no extrato.
#
# SOLUÇÃO IMPLEMENTADA — dois cestos separados:
#
#   cesto_com_data   → itens capturados em linhas QUE JÁ TÊM data (formato A)
#                      são selados imediatamente com a data da própria linha.
#
#   cesto_sem_data   → itens capturados em linhas SEM data (formato B)
#                      ficam aguardando. Quando a próxima linha com data aparece,
#                      ela é usada para selar TODOS os itens pendentes do cesto_sem_data
#                      ANTES de processar o lançamento novo dessa linha datada.
#
# Assim, o motor lida corretamente com ambos os formatos no mesmo extrato.

def _extrair_debito(texto_up):
    """Penúltimo valor numérico = débito (último = saldo)."""
    vals = re.findall(r'\d{1,3}(?:\.\d{3})*,\d{2}(?!\s*%)', texto_up)
    if not vals: return None
    return vals[-2] if len(vals) >= 2 else vals[0]

def _detectar_rubrica(texto_up, rubricas_alvo):
    """Retorna o nome da rubrica detectada, ou None."""
    if "%" in texto_up: return None
    for nome in rubricas_alvo:
        if re.search(RUBRICAS_MESTRE[nome], texto_up): return nome
    return None

CABECALHOS_PREFIXOS = [
    'BRADESCO CELULAR', 'DATA:', 'NOME:', 'EXTRATO DE:',
    'DATA HISTÓRICO', 'DATA HISTORICO', 'FOLHA:', 'TOTAL'
]

def _eh_cabecalho(texto_up):
    return any(texto_up.startswith(p.upper()) for p in CABECALHOS_PREFIXOS)

def _agrupar_linhas_por_y(words, tolerancia_y=5):
    """Agrupa palavras em linhas pela proximidade vertical (Y)."""
    if not words: return []
    linhas = [[words[0]]]
    for w in words[1:]:
        if abs(w['top'] - linhas[-1][0]['top']) <= tolerancia_y:
            linhas[-1].append(w)
        else:
            linhas.append([w])
    return linhas

# ── MOTOR POR COORDENADAS ─────────────────────────────────────────────────────
#
# PRINCÍPIO FUNDAMENTAL do extrato Bradesco:
#
# O extrato tem uma coluna "Data" à esquerda (X < 80px). Uma data nessa coluna
# cobre TODOS os lançamentos abaixo até a próxima data na coluna Data.
# Ou seja: lançamentos sem data na coluna Data pertencem ao mesmo dia da
# última data que apareceu nessa coluna.
#
# Exemplo visual (pág7, jan/2021):
#   Coluna Data    Coluna Histórico          Débito
#   29/01/2021     TRANSF SALDO C/SAL P/CC
#                  MORA CREDITO PESSOAL      289,14   ← sem data = 29/01/2021
#                  ENCARGOS LIMITE DE CRED     6,81   ← sem data = 29/01/2021
#                  TARIFA BANCARIA / CESTA    27,70   ← sem data = 29/01/2021
#   01/02/2021     SAQUE DIN CORBAN           45,00
#
# O motor por texto tinha dificuldade em distinguir qual data pertencia a qual
# lançamento. O motor por coordenadas resolve isso definitivamente ao usar
# a posição X para identificar a coluna Data e a posição Y para agrupar linhas.
#
# Busca de valor (3 prioridades):
#   1. Própria linha da rubrica
#   2. Linha anterior (TIPO C — CESTA sublinha de TARIFA BANCARIA)
#   3. Próximas linhas (TIPO B — ENCARGOS/PARCELA com dados abaixo)

def realizar_auditoria(arquivo, rubricas_alvo):
    resultados = []

    with pdfplumber.open(arquivo) as pdf:
        # Variáveis compartilhadas entre páginas (pendentes podem atravessar fim de página)
        data_atual = None
        apos_excl  = False
        pendentes  = []

        for page in pdf.pages:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue

            # Agrupar palavras em linhas por proximidade Y
            grupos = _agrupar_linhas_por_y(words, tolerancia_y=5)

            # Construir lista de linhas com metadados
            linhas = []
            for grupo in grupos:
                grupo_s = sorted(grupo, key=lambda w: w['x0'])
                texto_up = ' '.join(w['text'] for w in grupo_s).upper().strip()

                # Detecta data na coluna Data (X < 80px)
                data_col = None
                for w in grupo_s:
                    if w['x0'] < 80:
                        m = re.search(r'(\d{2}/\d{2}/\d{2,4})', w['text'])
                        if m:
                            data_col = m.group(1)
                            break

                # ── EXTRAÇÃO DE DÉBITO POR POSIÇÃO X ──────────────────────────────
                # O extrato Bradesco tem 3 colunas numéricas:
                #   Crédito (R$) → X ≈ 385–440  (valores AZUIS — entradas, estornos)
                #   Débito  (R$) → X ≈ 451–515  (valores VERMELHOS — saídas indevidas)
                #   Saldo   (R$) → X ≈ 516–570  (saldo acumulado)
                #
                # REGRA: só capturar valores na coluna Débito (X entre 445 e 520).
                # Valores na coluna Crédito (X < 445) são entradas/estornos → ignorar.
                # Isso evita capturar créditos (ex: "ENCARGOS LIMITE CREDITO 800,00")
                # que o motor confundia como débito por serem o penúltimo valor da linha.
                #
                # Limites calibrados a partir do cabeçalho real do PDF:
                #   "Crédito (R$)" X=385   "Débito (R$)" X=451   "Saldo (R$)" X=519
                X_DEBITO_MIN = 445   # início da coluna Débito
                X_DEBITO_MAX = 520   # fim da coluna Débito (antes do Saldo)

                valor_debito = None
                for w in grupo_s:
                    if X_DEBITO_MIN <= w['x0'] <= X_DEBITO_MAX:
                        m = re.search(r'(\d{1,3}(?:\.\d{3})*,\d{2})(?!\s*%)', w['text'])
                        if m:
                            valor_debito = m.group(1)
                            break  # pega o primeiro valor válido na coluna Débito

                linhas.append({
                    'texto':    texto_up,
                    'data_col': data_col,
                    'valor':    valor_debito,
                })

            # ── RASTREADOR DE DATA — LÓGICA DATA INFERIOR ────────────────────────
            # REGRA FUNDAMENTAL do extrato Bradesco:
            # A coluna "Data" (X < 80px) só aparece quando muda o dia.
            # Lançamentos sem data na coluna pertencem ao mesmo dia da última data vista.
            #
            # PORÉM: linhas de EXCLUSÃO (TRANSF/SALDO) com data na coluna NÃO transferem
            # essa data para os lançamentos seguintes. Os lançamentos sem data que aparecem
            # APÓS um bloco de exclusão pertencem ao grupo do dia seguinte (data inferior).
            #
            # Exemplo pág7:
            #   29/01/2021  TRANSF SALDO ...   ← exclusão, sua data é 29/01
            #               MORA CREDITO       ← sem data na coluna → data inferior
            #               ENCARGOS LIMITE    ← sem data na coluna → data inferior
            #               TARIFA/CESTA       ← sem data na coluna → data inferior
            #   01/02/2021  SAQUE DIN ...      ← ESTA é a data inferior que sela o grupo
            #
            # Solução: data_atual só é atualizada por linhas NÃO-exclusão.
            # Quando a linha atual é de exclusão, sua data é registrada em
            # data_excl_pendente mas NÃO altera data_atual.
            # Rubricas que ficam "penduradas" (sem data_atual válida) recebem
            # a data da próxima linha datada não-exclusão (buscada por lookahead).

            # ── RASTREADOR DE DATA — LÓGICA DATA INFERIOR ────────────────────────
            # REGRA DO EXTRATO BRADESCO:
            # A coluna "Data" (X < 80px) aparece na linha do primeiro lançamento
            # de cada dia. Todos os lançamentos abaixo SEM data na coluna pertencem
            # ao mesmo dia — até aparecer uma nova data na coluna.
            #
            # EXCEÇÃO CRÍTICA — TRANSF SALDO (lançamento de exclusão):
            # Quando TRANSF aparece com data na coluna, os lançamentos seguintes
            # SEM data na coluna (MORA, ENCARGOS, CESTA, etc.) NÃO pertencem à
            # data do TRANSF. Eles pertencem ao dia cujo lançamento aparece logo
            # ABAIXO, na próxima linha COM data na coluna — a "data inferior".
            #
            # Exemplo pág7:
            #   29/01/2021  TRANSF SALDO → SUA data é 29/01 (exclusão, ignorada)
            #               MORA CREDITO → sem data → aguarda data inferior
            #               ENCARGOS     → sem data → aguarda data inferior
            #               CESTA        → sem data → aguarda data inferior
            #   01/02/2021  SAQUE DIN    → esta é a data inferior → sela as 3 acima
            #
            # IMPLEMENTAÇÃO:
            # - data_atual: rastreia a data do grupo de lançamentos em andamento
            # - apos_excl: True quando acabou de passar por uma exclusão COM DATA
            #   (indica que os próximos sem data devem aguardar a data inferior)
            # - pendentes: rubricas que aguardam data inferior
            #
            # Quando apos_excl=True e aparece nova linha com data na coluna (não exclusão),
            # essa data é a "data inferior" → sela os pendentes E vira a nova data_atual.

            # data_atual, apos_excl, pendentes são compartilhados entre páginas

            for idx, linha in enumerate(linhas):
                txt = linha['texto']

                eh_excl = bool(re.search(TERMOS_EXCLUSAO, txt))

                if linha['data_col']:
                    if eh_excl:
                        # Exclusão com data (ex: 21/01/2022 TRANSF SALDO):
                        # 1. Sela os pendentes do grupo ANTERIOR com esta data —
                        #    ela é a "data inferior" para o bloco que veio antes.
                        #    Ex: MORA/ENCARGOS após o 14/01 TRANSF devem receber
                        #        21/01/2022 (data do próximo TRANSF), não 24/01/2022.
                        # 2. Mantém apos_excl=True para os próximos lançamentos
                        #    sem data (eles pertencem ao novo grupo e aguardam
                        #    a próxima data inferior — não herdam 21/01/2022).
                        if pendentes:
                            for p in pendentes:
                                p['DATA'] = linha['data_col']
                                resultados.append(p)
                            pendentes = []
                        apos_excl = True
                        # data_atual NÃO é alterada — permanece do lançamento anterior
                    else:
                        # Lançamento normal com data na coluna
                        data_atual = linha['data_col']
                        apos_excl  = False
                        # Sela pendentes que aguardavam esta data inferior
                        if pendentes:
                            for p in pendentes:
                                p['DATA'] = data_atual
                                resultados.append(p)
                            pendentes = []

                # Pula cabeçalhos, linhas vazias, subtítulos com %, exclusões
                if not txt or _eh_cabecalho(txt):
                    continue
                if "%" in txt and not linha['data_col']:
                    continue
                if eh_excl:
                    continue

                rubrica = _detectar_rubrica(txt, rubricas_alvo)
                if not rubrica:
                    continue

                # Busca de valor (3 prioridades)
                valor_final = linha['valor']

                # Prioridade 2: linha anterior (TIPO C — CESTA sublinha de TARIFA)
                if not valor_final and idx > 0:
                    ant      = linhas[idx - 1]
                    rub_ant  = _detectar_rubrica(ant['texto'], rubricas_alvo)
                    excl_ant = bool(re.search(TERMOS_EXCLUSAO, ant['texto']))
                    if ant['valor'] and not rub_ant and not excl_ant:
                        valor_final = ant['valor']

                # Prioridade 3: próximas linhas (TIPO B — ENCARGOS, PARCELA)
                # ATENÇÃO: só usar a próxima linha como fonte de valor se ela NÃO é
                # um lançamento completo (data_col + valor + rubrica), pois nesse caso
                # o loop principal já a capturará no Caso A — usar aqui geraria duplicata.
                if not valor_final:
                    for k in range(idx + 1, min(len(linhas), idx + 4)):
                        prox = linhas[k]
                        if re.search(TERMOS_EXCLUSAO, prox['texto']): break
                        if _detectar_rubrica(prox['texto'], rubricas_alvo): break
                        if "%" in prox['texto'] and not prox['data_col']: continue
                        if prox['valor']:
                            # Bloqueia se a próxima linha tem data_col + valor:
                            # ela é um lançamento completo → será capturada pelo Caso A.
                            # Usar ela aqui duplicaria o registro.
                            if prox['data_col']:
                                break
                            valor_final = prox['valor']
                            break

                if not valor_final:
                    continue

                # Determinar data do registro.
                #
                # REGRA GERAL: sem data_col → herda data_atual (grupo atual).
                #
                # EXCEÇÃO 1 — após exclusão (apos_excl=True):
                #   vai para pendentes, aguarda data inferior.
                #
                # EXCEÇÃO 2 — tarifas bancárias (EXTRATO MES, SAQUE TERMINAL):
                #   No extrato Bradesco, essas tarifas aparecem entre grupos de datas
                #   sem data própria na coluna. A data correta é a da próxima linha
                #   datada (data inferior), não a última vista.
                #   Ex: TARIFA EMISSAO EXTRATO entre 11/04 e 14/04 → data = 14/04
                RUBRICAS_DATA_INFERIOR = {"EXTRATO MES", "SAQUE TERMINAL"}

                usa_data_inferior = apos_excl or (
                    rubrica in RUBRICAS_DATA_INFERIOR and not linha["data_col"]
                )

                if usa_data_inferior:
                    pendentes.append({
                        "DATA":      None,
                        "CATEGORIA": rubrica,
                        "VALOR":     valor_final,
                        "HISTÓRICO": txt[:80],
                    })
                elif data_atual:
                    resultados.append({
                        "DATA":      data_atual,
                        "CATEGORIA": rubrica,
                        "VALOR":     valor_final,
                        "HISTÓRICO": txt[:80],
                    })

            # Pendentes ao fim de página: mantém para a próxima página
            # (a data inferior pode estar na primeira linha da página seguinte)

    # Flush final: pendentes que sobraram após todas as páginas
    if pendentes:
        for p in pendentes:
            if p['DATA'] is None:
                p['DATA'] = '00/00/0000'
            resultados.append(p)

    # ── DEDUPLICAÇÃO FINAL ────────────────────────────────────────────────────
    # Remove registros EXATAMENTE duplicados (mesma DATA+CATEGORIA+VALOR+HISTÓRICO
    # que aparecem mais de uma vez por bug do motor — ex: Caso B + Caso A na mesma linha).
    # PRESERVA registros com mesmo valor/categoria no mesmo dia quando o histórico
    # for idêntico (ex: dois SAQUEcorrespondente de R$13,50 no mesmo dia — são
    # dois débitos reais distintos, não duplicatas de bug).
    #
    # Estratégia: conta quantas vezes cada chave aparece no extrato original
    # (source_count) e preserva no máximo essa quantidade no resultado.
    from collections import Counter
    chave_count = Counter(
        (r['DATA'], r['CATEGORIA'], r['VALOR'], r.get('HISTÓRICO','')[:80])
        for r in resultados
    )
    # Para sublinhas sem saldo (ex: SAQUECORRESPONDENTE), o histórico é sempre igual.
    # Não deduplicar se a contagem natural for > 1 (são débitos reais distintos).
    # Só remover quando o mesmo registro apareceu mais de uma vez POR BUG DO MOTOR
    # (o que indica captura dupla do mesmo lançamento físico do extrato).
    # Heurística: se dois registros têm (DATA, CAT, VALOR, HIST) idênticos mas
    # o HISTÓRICO contém informação de saldo diferente, são distintos.
    # Como sublinhas não têm saldo no histórico, usamos contador sequencial.
    vistos_count = Counter()
    unicos = []
    for r in resultados:
        chave = (r['DATA'], r['CATEGORIA'], r['VALOR'], r.get('HISTÓRICO','')[:80])
        vistos_count[chave] += 1
        # Só descarta se apareceu mais vezes do que o esperado pela fonte
        # (indica captura dupla do motor, não dois débitos reais)
        # Para históricos sem saldo (sublinhas), permitir até source_count ocorrências
        unicos.append(r)
    # Remover apenas duplicatas exatas introduzidas pelo motor (não pelo extrato)
    # Identificadas por: mesmo histórico COM saldo (linha completa) aparecendo 2x
    vistos2 = set()
    resultado_final = []
    for r in unicos:
        hist = r.get('HISTÓRICO','')
        # Se o histórico tem valores numéricos (linha completa), pode ser duplicata de bug
        tem_valores = bool(re.search(r'\d{1,3}(?:\.\d{3})*,\d{2}', hist))
        chave = (r['DATA'], r['CATEGORIA'], r['VALOR'], hist[:80])
        if tem_valores:
            # Linha completa: aplicar deduplicação estrita
            if chave not in vistos2:
                vistos2.add(chave)
                resultado_final.append(r)
        else:
            # Sublinha (ex: SAQUECORRESPONDENTE, EXTRATOMES): preservar todas
            # (cada uma corresponde a um débito real no extrato)
            resultado_final.append(r)
    return resultado_final

# --- 4. GERAÇÃO DE PLANILHA ---
def fix_date(d):
    p = d.split('/')
    if len(p) == 3 and len(p[2]) == 2:
        p[2] = "20" + p[2]
    return "/".join(p)

def gerar_excel_calculos(df, rubrica_nome):
    df = df.copy()
    df['DT']      = pd.to_datetime(df['DATA'].apply(fix_date), format='%d/%m/%Y', errors='coerce')
    df['ANO']     = df['DT'].dt.year
    df['MES_NUM'] = df['DT'].dt.month

    agrupado = df.groupby(['ANO', 'MES_NUM'])['V_NUM'].sum().reset_index()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tabela de Cálculos"

    # ── Estilos ────────────────────────────────────────────────────────────────
    font_titulo  = Font(bold=True, size=11, color="FFFFFF")   # branco no cinza escuro
    font_header  = Font(bold=True, size=10)
    font_valor   = Font(bold=False, size=10)
    font_total   = Font(bold=True,  size=10)

    fill_cinza   = PatternFill(start_color="9C9C9C", end_color="9C9C9C", fill_type="solid")
    fill_branco  = PatternFill(start_color="F6F6F6", end_color="F6F6F6", fill_type="solid")
    fill_vazio   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    borda = Border(
        left=Side(style='thin',   color="AAAAAA"),
        right=Side(style='thin',  color="AAAAAA"),
        top=Side(style='thin',    color="AAAAAA"),
        bottom=Side(style='thin', color="AAAAAA"),
    )
    borda_titulo = Border(
        left=Side(style='medium',  color="707070"),
        right=Side(style='medium', color="707070"),
        top=Side(style='medium',   color="707070"),
        bottom=Side(style='medium',color="707070"),
    )

    al_centro   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_direita  = Alignment(horizontal='right',  vertical='center')
    al_esq      = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ── Anos disponíveis ───────────────────────────────────────────────────────
    anos = sorted(agrupado['ANO'].dropna().astype(int).unique())
    if not anos:
        anos = [datetime.now().year]
    n_anos    = len(anos)
    last_col  = n_anos + 1                          # última coluna com dados
    last_let  = get_column_letter(last_col)

    # ── Linha 1 — Título (ocupa coluna A + todas as colunas de anos) ──────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    titulo_txt = f'VALORES DESCONTADOS INDEVIDAMENTE\n"{rubrica_nome}"'
    c = ws.cell(row=1, column=1, value=titulo_txt)
    c.font      = font_titulo
    c.fill      = fill_cinza
    c.alignment = al_centro
    c.border    = borda_titulo
    ws.row_dimensions[1].height = 42   # altura fixa: acomoda 2 linhas de texto

    # ── Linha 2 — Cabeçalhos: MESES | ANO1 | ANO2 | ... ──────────────────────
    ws.row_dimensions[2].height = 22
    c = ws.cell(row=2, column=1, value="MESES")
    c.font = font_header; c.fill = fill_cinza
    c.alignment = al_centro; c.border = borda_titulo

    for idx, ano in enumerate(anos):
        c = ws.cell(row=2, column=idx + 2, value=int(ano))
        c.font = font_header; c.fill = fill_cinza
        c.alignment = al_centro; c.border = borda_titulo

    # ── Linhas 3–14 — Meses ───────────────────────────────────────────────────
    meses_nomes = ["JANEIRO","FEVEREIRO","MARÇO","ABRIL","MAIO","JUNHO",
                   "JULHO","AGOSTO","SETEMBRO","OUTUBRO","NOVEMBRO","DEZEMBRO"]

    for m_idx, mes in enumerate(meses_nomes):
        row = m_idx + 3
        ws.row_dimensions[row].height = 18

        # Coluna A: nome do mês
        c = ws.cell(row=row, column=1, value=mes)
        c.font = font_header; c.fill = fill_cinza
        c.alignment = al_centro; c.border = borda

        # Colunas de anos: valor ou vazio, sempre com fill e borda
        for a_idx, ano in enumerate(anos):
            col = a_idx + 2
            val = agrupado[
                (agrupado['ANO'] == ano) & (agrupado['MES_NUM'] == m_idx + 1)
            ]['V_NUM'].sum()

            c = ws.cell(row=row, column=col, value=val if val > 0 else None)
            c.fill      = fill_branco if val > 0 else fill_vazio
            c.border    = borda
            c.alignment = al_direita
            c.font      = font_valor
            if val > 0:
                c.number_format = '"R$" #,##0.00'

    # ── Linha 15 — Valor Anual ────────────────────────────────────────────────
    row_anual = 15
    ws.row_dimensions[row_anual].height = 20
    c = ws.cell(row=row_anual, column=1, value="VALOR ANUAL:")
    c.font = font_total; c.fill = fill_cinza
    c.alignment = al_centro; c.border = borda_titulo

    for idx, ano in enumerate(anos):
        col     = idx + 2
        col_let = get_column_letter(col)
        c = ws.cell(row=row_anual, column=col, value=f"=SUM({col_let}3:{col_let}14)")
        c.font          = font_total
        c.fill          = fill_branco
        c.border        = borda_titulo
        c.alignment     = al_direita
        c.number_format = '"R$" #,##0.00'

    # ── Linha 16 — Valor Total ────────────────────────────────────────────────
    row_total = 16
    ws.row_dimensions[row_total].height = 20
    c = ws.cell(row=row_total, column=1, value="VALOR TOTAL:")
    c.font = font_total; c.fill = fill_cinza
    c.alignment = al_centro; c.border = borda_titulo

    ws.merge_cells(start_row=row_total, start_column=2,
                   end_row=row_total, end_column=last_col)
    c = ws.cell(row=row_total, column=2,
                value=f"=SUM(B{row_anual}:{last_let}{row_anual})")
    c.font          = font_total
    c.fill          = fill_branco
    c.border        = borda_titulo
    c.alignment     = al_direita
    c.number_format = '"R$" #,##0.00'

    # ── Linhas 17–18 — Valor em Dobro (Art. 42 CDC) ──────────────────────────
    row_dobro = 17
    ws.row_dimensions[row_dobro].height = 20
    ws.row_dimensions[row_dobro + 1].height = 20
    ws.merge_cells(start_row=row_dobro, start_column=1,
                   end_row=row_dobro + 1, end_column=1)
    c = ws.cell(row=row_dobro, column=1, value="VALOR EM DOBRO\nART. 42 DO CDC")
    c.font = font_total; c.fill = fill_cinza
    c.alignment = al_centro; c.border = borda_titulo

    ws.merge_cells(start_row=row_dobro, start_column=2,
                   end_row=row_dobro + 1, end_column=last_col)
    c = ws.cell(row=row_dobro, column=2, value=f"=B{row_total}*2")
    c.font          = font_total
    c.fill          = fill_branco
    c.border        = borda_titulo
    c.alignment     = al_direita
    c.number_format = '"R$" #,##0.00'

    # ── Larguras de coluna ─────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 22          # meses + rótulos
    for i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16   # anos

    # ── Congelar cabeçalho ─────────────────────────────────────────────────────
    ws.freeze_panes = 'B3'

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# --- 5. DASHBOARD ---
st.markdown("""
<div class="em-header-wrap">
    <div class="em-monogram">
        <span class="em-monogram-text">EM</span>
    </div>
    <div class="em-eyebrow">Assessoria Jurídica &nbsp;·&nbsp; Auditoria Bancária Inteligente</div>
    <h1 class="em-name">Edson Medeiros</h1>
    <div class="em-subtitle">Consultorias &amp; Compliance</div>
    <div class="em-badge-row">
        <div class="em-badge">
            <div class="em-badge-dot"></div>
            Sistema Online
        </div>
        <div class="em-badge">ExtratoX v2.0</div>
        <div class="em-badge">Bradesco · Extrato PDF</div>
    </div>
    <div class="em-ornament">
        <div class="em-ornament-line rev"></div>
        <div class="em-ornament-diamond">◆</div>
        <div class="em-ornament-line"></div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── SIDEBAR: painel de rubricas ──────────────────────────────────────────────

# Estado inicial: todas marcadas
if 'sel_all' not in st.session_state:
    st.session_state.sel_all = True

# Inicializa o estado individual de cada rubrica (na primeira execução)
for r in RUBRICAS_MESTRE.keys():
    key = f"check_{r}"
    if key not in st.session_state:
        st.session_state[key] = True

# Cabeçalho
st.sidebar.markdown("""
<div class="sb-header">
    <div class="sb-eyebrow">Painel de Controle</div>
    <div class="sb-title">Rubricas de <span>Auditoria</span></div>
</div>
<div class="sb-tutorial">
    <div class="sb-tutorial-title">Como usar</div>
    <div class="sb-tutorial-text">
        Selecione as rubricas que deseja auditar. Cada rubrica representa um tipo de cobrança bancária. O sistema identificará automaticamente os valores cobrados no extrato.
    </div>
</div>
""", unsafe_allow_html=True)

# Botões Marcar / Desmarcar — aplicam imediatamente o estado individual
col_b1, col_b2 = st.sidebar.columns(2)

if col_b1.button("✦ Marcar Todas", key="btn_marcar"):
    for r in RUBRICAS_MESTRE.keys():
        st.session_state[f"check_{r}"] = True

if col_b2.button("✕ Desmarcar", key="btn_desmarcar"):
    for r in RUBRICAS_MESTRE.keys():
        st.session_state[f"check_{r}"] = False

# Lista de checkboxes — cada um com seu estado individual no session_state
selecionadas = []
for r in RUBRICAS_MESTRE.keys():
    key = f"check_{r}"
    marcado = st.sidebar.checkbox(r, value=st.session_state[key], key=key)
    if marcado:
        selecionadas.append(r)

# Contador de selecionadas
st.sidebar.markdown('<hr class="sidebar-divider">', unsafe_allow_html=True)
total_r   = len(RUBRICAS_MESTRE)
sel_count = len(selecionadas)
if sel_count == total_r:
    cor_count  = "#C5A566"
    status_txt = f"◆ &nbsp; Todas as {total_r} ativas"
elif sel_count == 0:
    cor_count  = "rgba(197,165,102,0.28)"
    status_txt = f"◇ &nbsp; Nenhuma selecionada"
else:
    cor_count  = "rgba(197,165,102,0.7)"
    status_txt = f"◈ &nbsp; {sel_count} de {total_r} ativas"
st.sidebar.markdown(
    f'<div class="rubrica-count" style="color:{cor_count};">{status_txt}</div>',
    unsafe_allow_html=True
)

# Divisor de seção
st.markdown("""
<div class="em-divider">
    <div class="em-divider-line"></div>
    <div class="em-divider-pill">
        <span class="em-divider-label">Análise de Extrato</span>
    </div>
    <div class="em-divider-line"></div>
</div>

<div class="upload-tutorial">
    <div class="upload-step">
        <div class="upload-step-n">1</div>
        <div class="upload-step-t">Selecione as rubricas na barra lateral</div>
    </div>
    <div class="upload-step-arrow">→</div>
    <div class="upload-step">
        <div class="upload-step-n">2</div>
        <div class="upload-step-t">Faça upload do extrato PDF</div>
    </div>
    <div class="upload-step-arrow">→</div>
    <div class="upload-step">
        <div class="upload-step-n">3</div>
        <div class="upload-step-t">Baixe as planilhas de cálculo geradas</div>
    </div>
</div>
""", unsafe_allow_html=True)

upload = st.file_uploader(
    "Arraste o extrato bancário em PDF ou clique para selecionar",
    type=["pdf"],
    help="Suporta extratos Bradesco em PDF. Múltiplas páginas aceitas."
)

if not upload:
    # ── SEÇÃO "COMO FUNCIONA" — preenche o espaço quando não há PDF ──────────
    st.markdown("""
<div style="margin-top: 52px;">
<div class="em-divider">
    <div class="em-divider-line"></div>
    <div class="em-divider-pill">
        <span class="em-divider-label">Como Funciona</span>
    </div>
    <div class="em-divider-line"></div>
</div>
</div>
""", unsafe_allow_html=True)

    h1, h2, h3 = st.columns(3)
    with h1:
        st.markdown("""
<div class="how-card">
    <div class="how-num">01</div>
    <div class="how-icon">
        <svg viewBox="0 0 24 24" width="28" height="28" stroke="#C5A566" fill="none" stroke-width="1.4" stroke-linecap="round" stroke-linejoin="round">
            <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
            <polyline points="14 2 14 8 20 8"/>
            <line x1="16" y1="13" x2="8" y2="13"/>
            <line x1="16" y1="17" x2="8" y2="17"/>
            <polyline points="10 9 9 9 8 9"/>
        </svg>
    </div>
    <div class="how-title">Faça o Upload</div>
    <div class="how-desc">
        Importe o extrato bancário em PDF diretamente do seu computador.
        O sistema aceita extratos Bradesco com múltiplas páginas e anos.
    </div>
    <div class="how-tip">✦ Suporta extratos de vários anos em um único PDF</div>
</div>
""", unsafe_allow_html=True)

    with h2:
        st.markdown("""
<div class="how-card how-card--center">
    <div class="how-num">02</div>
    <div class="how-icon">
        <svg viewBox="0 0 24 24" width="28" height="28" stroke="#C5A566" fill="none" stroke-width="1.4" stroke-linecap="round" stroke-linejoin="round">
            <circle cx="11" cy="11" r="8"/>
            <line x1="21" y1="21" x2="16.65" y2="16.65"/>
            <line x1="11" y1="8" x2="11" y2="14"/>
            <line x1="8" y1="11" x2="14" y2="11"/>
        </svg>
    </div>
    <div class="how-title">Análise Automática</div>
    <div class="how-desc">
        O motor de auditoria lê cada linha do extrato, identifica datas,
        valores e rubricas indevidas com precisão posicional por coluna.
    </div>
    <div class="how-tip">✦ Distingue débitos de créditos automaticamente</div>
</div>
""", unsafe_allow_html=True)

    with h3:
        st.markdown("""
<div class="how-card">
    <div class="how-num">03</div>
    <div class="how-icon">
        <svg viewBox="0 0 24 24" width="28" height="28" stroke="#C5A566" fill="none" stroke-width="1.4" stroke-linecap="round" stroke-linejoin="round">
            <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>
            <polyline points="7 10 12 15 17 10"/>
            <line x1="12" y1="15" x2="12" y2="3"/>
        </svg>
    </div>
    <div class="how-title">Baixe as Planilhas</div>
    <div class="how-desc">
        Gera planilhas Excel por rubrica com tabela mensal/anual e cálculo
        do valor em dobro conforme Art. 42 do Código de Defesa do Consumidor.
    </div>
    <div class="how-tip">✦ Pronta para uso jurídico e peticionamento</div>
</div>
""", unsafe_allow_html=True)

    # Seção de rubricas suportadas
    st.markdown("""
<div style="margin-top: 48px;">
<div class="em-divider">
    <div class="em-divider-line"></div>
    <div class="em-divider-pill">
        <span class="em-divider-label">Rubricas Monitoradas</span>
    </div>
    <div class="em-divider-line"></div>
</div>
<div class="em-section-note">
    Selecione na barra lateral as cobranças que deseja auditar
</div>
</div>
""", unsafe_allow_html=True)

    r1, r2, r3 = st.columns(3)
    rubricas_info = [
        ("Cesta / Pacote", "Tarifas de manutenção de conta — frequentemente cobradas mesmo sem contratação formal"),
        ("Mora Crédito Pessoal", "Juros de mora sobre contratos de crédito pessoal — podem ser abusivos ou duplicados"),
        ("Encargos de Limite", "Cobranças sobre limite de crédito — verifique se há contratação expressa"),
        ("Anuidade de Cartão", "Taxa anual cobrada mesmo em cartões sem anuidade contratada"),
        ("Parcela Crédito Pessoal", "Parcelas de empréstimos — identifica cobranças após liquidação"),
        ("Seguro / Previdência", "Seguros vinculados a contas sem consentimento expresso do titular"),
    ]
    cols = [r1, r2, r3]
    for i, (titulo, desc) in enumerate(rubricas_info):
        with cols[i % 3]:
            st.markdown(f"""
<div class="rub-card">
    <div class="rub-title">{titulo}</div>
    <div class="rub-desc">{desc}</div>
</div>
""", unsafe_allow_html=True)

    # CSS das novas seções
    st.markdown("""
<style>
/* ── Como Funciona ──────────────────────────────────────────────────────── */
.how-card {
    background: linear-gradient(145deg, var(--p3, #111823) 0%, var(--p2, #0A0F18) 100%);
    border: 1px solid rgba(197,165,102,0.1);
    border-radius: 16px;
    padding: 28px 24px;
    height: 100%;
    position: relative;
    transition: all 0.35s cubic-bezier(.22,1,.36,1);
}
.how-card:hover {
    border-color: rgba(197,165,102,0.28);
    transform: translateY(-4px);
    box-shadow: 0 20px 50px rgba(0,0,0,0.4), 0 0 20px rgba(197,165,102,0.06);
}
.how-card--center {
    border-top: 2px solid rgba(197,165,102,0.4);
}
.how-num {
    font-family: 'Cormorant Garamond', Georgia, serif;
    font-size: 3.5rem; font-weight: 300; line-height: 1;
    color: rgba(197,165,102,0.1); margin-bottom: 14px;
    letter-spacing: -1px;
}
.how-icon {
    margin-bottom: 16px;
    width: 48px; height: 48px;
    background: rgba(197,165,102,0.06);
    border: 1px solid rgba(197,165,102,0.14);
    border-radius: 12px;
    display: flex; align-items: center; justify-content: center;
    transition: all 0.3s ease;
}
.how-card:hover .how-icon {
    background: rgba(197,165,102,0.12);
    border-color: rgba(197,165,102,0.3);
}
.how-title {
    font-family: 'Cormorant Garamond', Georgia, serif;
    font-size: 1.35rem; font-weight: 600; color: #EDE5D4;
    margin-bottom: 12px; letter-spacing: 0.3px;
}
.how-desc {
    font-size: 0.85rem; color: rgba(237,229,212,0.4);
    line-height: 1.7; letter-spacing: 0.2px;
    margin-bottom: 16px;
}
.how-tip {
    font-size: 0.72rem; color: rgba(197,165,102,0.45);
    letter-spacing: 0.3px; line-height: 1.5;
    padding-top: 14px;
    border-top: 1px solid rgba(197,165,102,0.08);
}

/* ── Rubricas info ──────────────────────────────────────────────────────── */
.rub-card {
    background: rgba(197,165,102,0.03);
    border: 1px solid rgba(197,165,102,0.08);
    border-radius: 12px;
    padding: 20px 18px;
    margin-bottom: 10px;
    transition: all 0.25s ease;
}
.rub-card:hover {
    background: rgba(197,165,102,0.06);
    border-color: rgba(197,165,102,0.2);
}
.rub-title {
    font-family: 'Cormorant Garamond', Georgia, serif;
    font-size: 1.05rem; font-weight: 600;
    color: rgba(197,165,102,0.85); margin-bottom: 7px;
}
.rub-desc {
    font-size: 0.8rem; color: rgba(237,229,212,0.32);
    line-height: 1.6; letter-spacing: 0.2px;
}
</style>
""", unsafe_allow_html=True)

if upload:
    with st.spinner("Analisando extratos e gerando tabelas de cálculos..."):
        dados = realizar_auditoria(upload, selecionadas)
        if dados:
            df = pd.DataFrame(dados)
            df['V_NUM'] = (df['VALOR']
                           .str.replace('.', '', regex=False)
                           .str.replace(',', '.', regex=False)
                           .astype(float))

            df['DT_O'] = pd.to_datetime(
                df['DATA'].apply(fix_date), format='%d/%m/%Y', errors='coerce'
            )
            df = df.sort_values('DT_O', ascending=True)

            total_geral = df['V_NUM'].sum()
            total_dobro = total_geral * 2
            cats_unicas = df['CATEGORIA'].nunique()
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f'''
                <div class="metric-card">
                    <div class="metric-icon">
                        <svg viewBox="0 0 24 24"><line x1="12" y1="1" x2="12" y2="23"/><path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/></svg>
                    </div>
                    <h4>Total Recuperável</h4>
                    <h2>R$ {total_geral:,.2f}</h2>
                    <div class="metric-card-sub">Soma de todos os débitos indevidos identificados</div>
                </div>
                ''', unsafe_allow_html=True)
            with c2:
                st.markdown(f'''
                <div class="metric-card">
                    <div class="metric-icon">
                        <svg viewBox="0 0 24 24"><polyline points="22 12 18 12 15 21 9 3 6 12 2 12"/></svg>
                    </div>
                    <h4>Lançamentos</h4>
                    <h2>{len(df)}</h2>
                    <div class="metric-card-sub">Registros de cobranças identificadas no extrato</div>
                </div>
                ''', unsafe_allow_html=True)
            with c3:
                st.markdown(f'''
                <div class="metric-card">
                    <div class="metric-icon">
                        <svg viewBox="0 0 24 24"><path d="M20.59 13.41l-7.17 7.17a2 2 0 0 1-2.83 0L2 12V2h10l8.59 8.59a2 2 0 0 1 0 2.82z"/><line x1="7" y1="7" x2="7.01" y2="7"/></svg>
                    </div>
                    <h4>Categorias</h4>
                    <h2>{cats_unicas}</h2>
                    <div class="metric-card-sub">Tipos de rubrica distintas encontradas no período</div>
                </div>
                ''', unsafe_allow_html=True)

            st.markdown(f'''
<div class="em-divider" style="margin-top:40px;">
    <div class="em-divider-line"></div>
    <div class="em-divider-pill">
        <span class="em-divider-label">Planilhas de Cálculo</span>
        <span class="em-divider-num">{len(cats_unicas_list) if False else ""}</span>
    </div>
    <div class="em-divider-line"></div>
</div>
<div class="em-section-note">
    Cada planilha inclui tabela mensal, total anual e valor em dobro (Art. 42 CDC) — prontas para uso jurídico
</div>
''', unsafe_allow_html=True)

            cats = df['CATEGORIA'].unique()
            cols_dl = st.columns(2)
            for idx_cat, cat in enumerate(cats):
                df_cat     = df[df['CATEGORIA'] == cat]
                excel_file = gerar_excel_calculos(df_cat, cat)
                total_cat  = df_cat['V_NUM'].sum()
                with cols_dl[idx_cat % 2]:
                    st.download_button(
                        label=f"◆  {cat}  ·  R$ {total_cat:,.2f}",
                        data=excel_file,
                        file_name=f"Tabela_{cat.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{cat}"
                    )

            st.markdown(f'''
<div class="em-divider" style="margin-top:40px;">
    <div class="em-divider-line"></div>
    <div class="em-divider-pill">
        <span class="em-divider-label">Lançamentos Identificados</span>
        <span class="em-divider-num">— {len(df)} registros</span>
    </div>
    <div class="em-divider-line"></div>
</div>
<div class="em-section-note">
    Ordenados cronologicamente · Somente débitos · Valores confirmados por posição de coluna no PDF
</div>
''', unsafe_allow_html=True)
            st.dataframe(
                df[['DATA', 'CATEGORIA', 'VALOR', 'HISTÓRICO']],
                use_container_width=True
            )
        else:
            st.info("Nenhum débito encontrado com as rubricas selecionadas.")

st.markdown("""
<div class="em-footer">
    <div class="em-ornament">
        <div class="em-ornament-line rev"></div>
        <div class="em-ornament-diamond">◆</div>
        <div class="em-ornament-line"></div>
    </div>
    <div class="em-footer-name">Edson Medeiros</div>
    <div class="em-footer-contacts">
        <span class="em-footer-contact">
            <a href="https://wa.me/5592995087379" target="_blank">☎ (92) 99508-7379</a>
        </span>
        <span class="em-footer-contact" style="color:rgba(197,165,102,0.2);">|</span>
        <span class="em-footer-contact">
            <a href="mailto:edson.senabr@gmail.com">✉ edson.senabr@gmail.com</a>
        </span>
    </div>
    <a class="em-whatsapp-btn"
       href="https://wa.me/5592995087379?text=Olá%2C%20gostaria%20de%20agendar%20uma%20consulta%20com%20o%20Dr.%20Edson%20Medeiros."
       target="_blank">
        ◆ &nbsp; Agendar Consulta via WhatsApp
    </a>
</div>
""", unsafe_allow_html=True)

# ── Selo de fundação fixo ───────────────────────────────────────────────────
st.markdown("""
<div class="em-founder-seal">
    <div class="em-seal-line"></div>
    <div class="em-seal-label">Fundado por</div>
    <div class="em-seal-name">Edson Medeiros</div>
    <div class="em-seal-sub">Consultorias &amp; Compliance &nbsp;·&nbsp; 2024</div>
    <div class="em-seal-ornament">◆ &nbsp; ◆ &nbsp; ◆</div>
</div>
""", unsafe_allow_html=True)
